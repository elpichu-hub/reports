def seconds_to_hms(seconds):
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = int(seconds % 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


# Function to convert time string to seconds
def time_to_seconds(time_str):
    time_parts = time_str.split(':')
    hours = int(time_parts[0])
    minutes = int(time_parts[1])
    seconds = int(time_parts[2])
    return hours * 3600 + minutes * 60 + seconds


date_string = '4.26.2023'
file_path1 = 'User Productivity Summary stats two call center 5.15.2023 teams.csv'
def run_report(date_string, file_path1):
    import os
    import pandas as pd
    import openpyxl
    from datetime import datetime, time
    import calendar
    from openpyxl import utils, styles
    from openpyxl.styles import PatternFill


    # Specify the path to the Excel file
    output_path = f"{date_string} Daily Stats.xlsx"
    
    print(date_string, file_path1)
    # read the CSV file into a pandas dataframe
    df = pd.read_csv(file_path1)   

    # specify the column names you want to keep
    user_id = 'i3user'
    acd_calls = 'totEnteredACD' # calls offered
    acd_answered = 'totnAnsweredACD' # calls answered
    total_hold = 'totTHoldACD'
    total_acw = 'totTACW'
    tot_talk = 'totTTalkACD'
    full_name = 'DisplayUserName'
    calls_transfered = 'totnTransferedACD'
    team = 'JobTitle'
    
    # keep only selected columns in the dataframe
    df_custom = df[[user_id, acd_calls, acd_answered, total_hold, total_acw, tot_talk, full_name, calls_transfered, team]]
    print(df_custom)

    # Get the unique team values
    unique_teams = df_custom[team].unique()

    # Iterate over the unique teams
    for team_name in unique_teams:
        # Filter the DataFrame to keep only rows for the current team
        team_df = df_custom[df_custom[team] == team_name]

        

        # group the data by the 'i3user' column and calculate the total 'totEnteredACD',   'totTHoldACD', 'totTACW', and 'totTTalkACD'
        df_grouped = team_df.groupby(user_id).agg({
            acd_calls: 'sum',
            acd_answered: 'sum',
            total_hold: 'sum',
            total_acw: 'sum',
            tot_talk: 'sum',
            calls_transfered: 'sum',
            full_name: 'first',
        }).reset_index()
        
        # # Get the date from user input
        # # date = input("Enter the date in yyyy-mm-dd format: ")


        if df_grouped.isna().any().any():
            df_grouped = df_grouped.fillna(0)
            print("Warning: The grouped data contains NaN values.")

        
        epsilon = 1e-8  # Small constant to avoid division by zero

        df_grouped['average_hold'] = (df_grouped[total_hold] / (df_grouped[acd_answered] + epsilon))
        df_grouped['total_hold'] = df_grouped[total_hold]

        df_grouped['average_acw'] = (df_grouped[total_acw] / (df_grouped[acd_answered] + epsilon))
        df_grouped['total_acw'] = df_grouped[total_acw]

        df_grouped['AHT'] = ((df_grouped[tot_talk] + df_grouped[total_acw]) / (df_grouped[acd_answered] + epsilon))

        df_grouped['calls_offered'] = df_grouped[acd_calls]
        df_grouped['calls_answered'] = df_grouped[acd_answered]
        df_grouped['calls_transfered'] = df_grouped[calls_transfered]

        df_grouped['total_talk'] = df_grouped[tot_talk]
        df_grouped['average_talk'] = (df_grouped[tot_talk] / df_grouped[acd_answered] + epsilon)


        # Remove rows where ACD Calls is 0
        df_grouped = df_grouped[df_grouped['calls_answered'] != 0]

        # Remove extra spaces in FullName column
        df_grouped['FullName'] = df_grouped['DisplayUserName'].str.strip().str.replace(r'\s+', ' ', regex=True)

        # Sort the dataframe based on the 'FullName' column (A to Z)
        df_grouped = df_grouped.sort_values(by='FullName')

        
        # # Reorder the columns as required
        columns = ['FullName', 'calls_offered', 'calls_answered', 'calls_transfered', 'total_talk', 'average_talk', 'total_hold', 'average_hold', 'total_acw', 'average_acw', 'AHT']
        df_grouped = df_grouped[columns]


        # Apply the helper function to the relevant columns
        for column in ['total_talk', 'average_talk', 'total_hold', 'average_hold', 'total_acw', 'average_acw', 'AHT']:
            df_grouped[column] = df_grouped[column].apply(seconds_to_hms)

        # Reorder the columns as required
        columns = ['FullName', 'calls_offered', 'calls_answered', 'calls_transfered', 'total_talk', 'average_talk', 'total_hold', 'average_hold', 'total_acw', 'average_acw', 'AHT']
        df_grouped = df_grouped[columns]

        # get the totals of all the values to add to daily stats
        total_calls_offered = df_grouped['calls_offered'].sum()
        total_calls_answered = df_grouped['calls_answered'].sum()
        total_calls_transfered = df_grouped['calls_transfered'].sum()

        total_total_talk = df_grouped['total_talk'].apply(time_to_seconds).sum()
        total_total_hold = df_grouped['total_hold'].apply(time_to_seconds).sum()
        total_total_acw = df_grouped['total_acw'].apply(time_to_seconds).sum()

        # Calculate the totals and format them back to time strings
        total_total_talk = seconds_to_hms(total_total_talk)
        total_total_hold = seconds_to_hms(total_total_hold)
        total_total_acw = seconds_to_hms(total_total_acw)

        # get the averages of all the values that are average based for the daily stats
        average_average_talk = df_grouped['average_talk'].apply(time_to_seconds).mean()
        average_average_hold = df_grouped['average_hold'].apply(time_to_seconds).mean()
        average_average_acw = df_grouped['average_acw'].apply(time_to_seconds).mean()
        average_AHT = df_grouped['AHT'].apply(time_to_seconds).mean()

        # Print the totals and averages
        print("Total Calls Offered:", total_calls_offered)
        print("Total Calls Answered:", total_calls_answered)
        print("Total Calls Transferred:", total_calls_transfered)
        print("Total Total Talk:", total_total_talk)
        print("Total Total Hold:", total_total_hold)
        print("Total Total ACW:", total_total_acw)
        print("Average Average Talk:", seconds_to_hms(int(average_average_talk)))
        print("Average Average Hold:", seconds_to_hms(int(average_average_hold)))
        print("Average Average ACW:", seconds_to_hms(int(average_average_acw)))
        print("Average AHT:", seconds_to_hms(int(average_AHT)))
        


        
        # Create the output path for the current team
        output_path = f"{date_string} Daily Stats - {team_name}.xlsx"
        # Save the DataFrame to an Excel sheet for the current team
        df_grouped.to_excel(output_path, index=False)





        # Open the file
        wb = openpyxl.load_workbook(output_path)

        # Get the active sheet
        sheet = wb.active

        # Set the width of all columns to 10 
        for column in range(1, sheet.max_column + 1):
            column_letter = utils.get_column_letter(column)
            sheet.column_dimensions[column_letter].width = 15

        # Insert a row at the top and at row 3
        sheet.insert_rows(2)
        sheet.insert_rows(0)
        # Adjust the width of column A
        sheet.column_dimensions['A'].width = 37
        # Merge and center cells A1 and A2
        sheet.merge_cells('A1:A2')
        merged_cell = sheet['A1']
        merged_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # Get the day of the week
        date_obj = datetime.strptime(date_string, '%m.%d.%Y')
        day_of_week = calendar.day_name[date_obj.weekday()]

        # Add value to cell A1
        merged_cell.value = f"Daily Stats {day_of_week} {date_string.replace('.', '/')}"


        # Center all cells in the document
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')


        # Set the values for cells B1, C1, and D1
        sheet['B1'].value = "Offered"
        sheet['C1'].value = "Answered"
        sheet['D1'].value = "Transferred"

        sheet['E1'].value = "Talk Time"
        sheet['G1'].value = "Hold Time"
        sheet['I1'].value = "ACW Time"
        sheet['K1'].value = "Handle Time"

        sheet['B2'].value = "#"
        sheet['C2'].value = "#"
        sheet['D2'].value = "#"
        sheet['E2'].value = "Duration"
        sheet['F2'].value = "Average"
        sheet['G2'].value = "Duration"
        sheet['H2'].value = "Average"
        sheet['I2'].value = "Duration"
        sheet['J2'].value = "Average"
        sheet['K2'].value = "Average"

        sheet['A3'].value = "Agents"
        sheet['B3'].value = total_calls_offered
        sheet['C3'].value = total_calls_answered
        sheet['D3'].value = total_calls_transfered
        sheet['E3'].value = total_total_talk
        sheet['F3'].value = seconds_to_hms(average_average_talk)
        sheet['G3'].value = total_total_hold
        sheet['H3'].value = seconds_to_hms(average_average_hold)
        sheet['I3'].value = total_total_acw
        sheet['J3'].value = seconds_to_hms(average_average_acw)
        sheet['K3'].value = seconds_to_hms(average_AHT)
        

        # Merge cells E1 and F1
        sheet.merge_cells('E1:F1')

        # Merge cells G1 and H1
        sheet.merge_cells('G1:H1')

        # Merge cells I1 and J1
        sheet.merge_cells('I1:J1')

        # Get cell A1
        cell_a1 = sheet['A1']
        # Set the font style to bold
        cell_a1.font = styles.Font(bold=True)
        
        # Colors for background
        blue = PatternFill(start_color='0ACDFF', end_color='0ACDFF', fill_type='solid')
        green = PatternFill(start_color='59FFA0', end_color='59FFA0', fill_type='solid')
        orange = PatternFill(start_color='FF7F11', end_color='FF7F11', fill_type='solid')
        red = PatternFill(start_color='C1292E', end_color='C1292E', fill_type='solid')
        grey = PatternFill(start_color='808080', end_color='808080', fill_type='solid')


        # Set the background color of rows 1 and 2 from column A to column K
        for row in sheet.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.fill = blue
                cell.font = styles.Font(bold=True)

        # Set the background color of row 3 to grey
        for cell in sheet[3]:
            cell.fill = styles.PatternFill(start_color='808080', end_color='808080', fill_type='solid')

        # Iterate over all cells in the worksheet
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = styles.Border(
                    left=styles.Side(style='thin'),
                    right=styles.Side(style='thin'),
                    top=styles.Side(style='thin'),
                    bottom=styles.Side(style='thin')
                )


        # Set the height of rows 1 and 2 to 25 and all other rows to 17
        for row in range(1, sheet.max_row + 1):
            if row == 1 or row == 2:
                sheet.row_dimensions[row].height = 25
                for cell in sheet[row]:
                    cell.font = styles.Font(size=12, bold=True)  # explicitly setting the font size for rows 1 and 2 to 12
            else:
                sheet.row_dimensions[row].height = 17
                for cell in sheet[row]:
                    cell.font = styles.Font(size=12)

        # on column K color based on conditions
        for cell in sheet['K']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                # if 6:30 or less green
                if isinstance(time_obj, time) and time_obj < time(hour=0, minute=6, second=30):
                    cell.fill = green
                if isinstance(time_obj, time) and time_obj <= time(hour=0, minute=7, second=0) and time_obj > time(hour=0, minute=6, second=30):
                    cell.fill = orange
                if isinstance(time_obj, time) and time_obj > time(hour=0, minute=7, second=0):
                    cell.fill = red
            except ValueError:
                pass

        # more than 20 seconds will be red
        for cell in sheet['J']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                if isinstance(time_obj, time) and time_obj > time(hour=0, minute=0, second=20):
                    cell.fill = red
                else:
                    cell.fill = green
            except ValueError:
                pass

        # Get the cell J3 and its value
        j3_cell = sheet['J3']
        value = j3_cell.value

        # if more than 30
        time_obj = datetime.strptime(str(j3_cell.value), "%H:%M:%S").time()
        if isinstance(time_obj, time) and time_obj > time(hour=0, minute=0, second=30):
            j3_cell.fill = red
        else:
            j3_cell.fill = green

        # more than 30 seconds will be red
        for cell in sheet['H']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                if isinstance(time_obj, time) and time_obj > time(hour=0, minute=0, second=30):
                    cell.fill = red
            except ValueError:
                pass

        # more than 7 minutes will be red
        for cell in sheet['F']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                if isinstance(time_obj, time) and time_obj > time(hour=0, minute=7, second=0):
                    cell.fill = red
            except ValueError:
                pass

        # Save the workbook
        wb.save(output_path)
    os.startfile(output_path)




















run_report(date_string, file_path1)

